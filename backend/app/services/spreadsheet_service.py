"""
EZ-NEXUS AI — Spreadsheet Service
Creates, formats, and exports Excel workbooks and CSV files.
Uses openpyxl for rich formatting; falls back to plain CSV if not installed.
"""

from __future__ import annotations
import csv
import io
import logging
from pathlib import Path
from typing import Any, Dict, List, Optional

logger = logging.getLogger(__name__)

# Column definitions for each workbook type
PATIENT_COLUMNS = [
    "patient_first_name", "patient_last_name", "date_of_birth", "phone", "email",
    "address", "city", "state", "zip_code",
    "payer_name", "insurance_member_id", "group_number",
    "equipment_requested", "diagnosis_codes", "prescribing_doctor",
    "status", "notes",
]

BOOKKEEPING_COLUMNS = [
    "transaction_date", "vendor_or_customer", "entry_type", "category",
    "amount", "tax_flag", "memo", "status",
]

INSURANCE_COLUMNS = [
    "patient_first_name", "patient_last_name", "date_of_birth",
    "payer_name", "plan_type", "member_id", "group_number",
    "policy_holder", "relationship_to_patient", "phone_number", "eligibility_status",
]

EQUIPMENT_COLUMNS = [
    "patient_name", "equipment_type", "diagnosis_or_reason",
    "prescribing_provider", "hcpcs_codes",
    "insurance_required", "prior_auth_required", "status", "notes",
]


class SpreadsheetService:
    """Creates professional Excel workbooks and CSV exports."""

    # ── Patient Workbook ──────────────────────────────────────────────────────

    def create_patient_workbook(self, rows: List[Dict[str, Any]], output_path: str) -> str:
        return self._create_workbook(rows, PATIENT_COLUMNS, output_path, "Patient Data", "#1e3a5f")

    # ── Bookkeeping Workbook ──────────────────────────────────────────────────

    def create_bookkeeping_workbook(self, rows: List[Dict[str, Any]], output_path: str) -> str:
        return self._create_workbook(rows, BOOKKEEPING_COLUMNS, output_path, "Bookkeeping", "#1a3a1a")

    # ── Insurance Workbook ────────────────────────────────────────────────────

    def create_insurance_workbook(self, rows: List[Dict[str, Any]], output_path: str) -> str:
        return self._create_workbook(rows, INSURANCE_COLUMNS, output_path, "Insurance Profiles", "#3a1a1a")

    # ── Equipment Workbook ────────────────────────────────────────────────────

    def create_equipment_workbook(self, rows: List[Dict[str, Any]], output_path: str) -> str:
        return self._create_workbook(rows, EQUIPMENT_COLUMNS, output_path, "Equipment Requests", "#1a2a3a")

    # ── Generic workbook builder ──────────────────────────────────────────────

    def _create_workbook(
        self,
        rows: List[Dict[str, Any]],
        columns: List[str],
        output_path: str,
        sheet_name: str,
        header_color: str = "#1e3a5f",
    ) -> str:
        try:
            import pandas as pd
            from openpyxl import load_workbook
            from openpyxl.styles import Font, PatternFill, Alignment, Border, Side

            import pandas as pd

            df = pd.DataFrame(rows)
            # Ensure all expected columns exist
            for col in columns:
                if col not in df.columns:
                    df[col] = ""
            # Reorder to match schema; keep any extra columns at the end
            extra_cols = [c for c in df.columns if c not in columns]
            df = df[columns + extra_cols]

            Path(output_path).parent.mkdir(parents=True, exist_ok=True)

            with pd.ExcelWriter(output_path, engine="openpyxl") as writer:
                df.to_excel(writer, index=False, sheet_name=sheet_name)

            # Apply professional formatting
            wb = load_workbook(output_path)
            ws = wb[sheet_name]

            header_fill = PatternFill("solid", fgColor=header_color.lstrip("#"))
            thin_border = Border(
                left=Side(style="thin"), right=Side(style="thin"),
                top=Side(style="thin"), bottom=Side(style="thin"),
            )

            for cell in ws[1]:
                cell.font = Font(bold=True, color="FFFFFF", size=11)
                cell.fill = header_fill
                cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
                cell.border = thin_border

            for row in ws.iter_rows(min_row=2):
                for cell in row:
                    cell.border = thin_border
                    cell.alignment = Alignment(vertical="top", wrap_text=True)

            # Auto-fit column widths (max 40 chars)
            for col in ws.columns:
                max_len = max((len(str(cell.value or "")) for cell in col), default=10)
                ws.column_dimensions[col[0].column_letter].width = min(max_len + 4, 40)

            ws.freeze_panes = "A2"
            ws.auto_filter.ref = ws.dimensions

            wb.save(output_path)
            logger.info("Workbook created: %s (%d rows)", output_path, len(rows))
            return output_path

        except ImportError:
            logger.warning("pandas/openpyxl not installed — falling back to CSV")
            return self._fallback_csv(rows, columns, output_path.replace(".xlsx", ".csv"))

    # ── CSV fallback ──────────────────────────────────────────────────────────

    def _fallback_csv(
        self,
        rows: List[Dict[str, Any]],
        columns: List[str],
        output_path: str,
    ) -> str:
        Path(output_path).parent.mkdir(parents=True, exist_ok=True)
        with open(output_path, "w", newline="", encoding="utf-8") as f:
            writer = csv.DictWriter(f, fieldnames=columns, extrasaction="ignore")
            writer.writeheader()
            writer.writerows(rows)
        return output_path

    # ── QuickBooks CSV template ───────────────────────────────────────────────

    def create_quickbooks_template(self, entries: List[Dict[str, Any]], output_path: str) -> str:
        """Generate a QuickBooks-compatible IIF/CSV import template."""
        qb_columns = [
            "Date", "Description", "Amount", "Account", "Vendor/Customer", "Memo", "Category",
        ]
        rows = []
        for e in entries:
            rows.append({
                "Date":             e.get("transaction_date", ""),
                "Description":      e.get("entry_type", ""),
                "Amount":           e.get("amount", 0),
                "Account":          "Accounts Payable" if e.get("entry_type") == "expense" else "Accounts Receivable",
                "Vendor/Customer":  e.get("vendor_or_customer", ""),
                "Memo":             e.get("memo", "")[:100],
                "Category":         e.get("category", "Uncategorized"),
            })
        return self._fallback_csv(rows, qb_columns, output_path)

    # ── In-memory CSV (for API streaming) ────────────────────────────────────

    def to_csv_bytes(self, rows: List[Dict[str, Any]], columns: List[str]) -> bytes:
        """Return CSV as bytes — useful for streaming responses."""
        buf = io.StringIO()
        writer = csv.DictWriter(buf, fieldnames=columns, extrasaction="ignore")
        writer.writeheader()
        writer.writerows(rows)
        return buf.getvalue().encode("utf-8")
